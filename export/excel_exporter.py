"""
export/excel_exporter.py
------------------------
Exporta mensajes a Excel (.xlsx) con encabezados y filas alternadas.
"""

from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from database.models import Mensaje

COLUMNAS_DEFAULT = [
    "fecha",
    "hora",
    "remitente",
    "texto",
    "equipo",
    "tipo_mensaje",
    "adjuntos",
    "estado_proceso",
]


def exportar_excel(
    mensajes: Iterable[Mensaje],
    ruta: Path,
    columnas: Optional[list[str]] = None,
) -> None:
    """Exporta los mensajes a un archivo .xlsx."""
    columnas = columnas or COLUMNAS_DEFAULT
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "mensajes"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E2E8E0", end_color="E2E8E0", fill_type="solid")
    alt_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

    ws.append(columnas)
    for col_idx in range(1, len(columnas) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for i, m in enumerate(mensajes, start=2):
        fila = _mensaje_a_fila(m, columnas)
        ws.append(fila)
        if i % 2 == 0:
            for col_idx in range(1, len(columnas) + 1):
                ws.cell(row=i, column=col_idx).fill = alt_fill

    wb.save(ruta)


def _mensaje_a_fila(m: Mensaje, columnas: list[str]) -> list[str]:
    """Convierte un Mensaje en una fila según las columnas definidas."""
    salida: list[str] = []
    for col in columnas:
        if col == "adjuntos":
            salida.append("|".join(m.adjuntos) if m.adjuntos else "")
        else:
            valor = getattr(m, col, "")
            salida.append("" if valor is None else str(valor))
    return salida
